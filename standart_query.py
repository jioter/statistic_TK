import os.path

import timeit

import openpyxl

# import psycopg2
# database = psycopg2.connect(database="test", user="user1", password="password", host="localhost", port="5432")
# cursor = database.cursor()

import sqlite3
conn = sqlite3.connect('statistic_db.db')
cursor = conn.cursor()


def query(col_name):
    """
    Select poll question and count number of answers.
    :param col_name:
    :return: certain question and sum of answers
    """
    # if no such file -> create it
    if not os.path.isfile("data/statistics_calculated.xlsx"):
        book = openpyxl.Workbook()
        book.save("data/statistics_calculated.xlsx")

    # separate query for column 8
    if col_name == 'col_8':
        get_col_name_8 = """SELECT %s FROM test WHERE id_count=2""" % col_name
        cursor.execute(get_col_name_8)
        col_name_8 = [item for t in cursor.fetchall() for item in t]

        book = openpyxl.load_workbook("data/statistics_calculated.xlsx")
        sheet = book.active
        sheet.append(col_name_8)
        book.save("data/statistics_calculated.xlsx")

        query8 = """SELECT DISTINCT(col_8) FROM test WHERE col_8 <> '' AND id_count > 3 AND col_8 NOT IN ('-') """
        cursor.execute(query8)
        records8 = cursor.fetchall()
        records8 = [item for t in records8 for item in t]
        records8 = [el.replace("'", "''") for el in records8]

        # split elements in sublists
        rez = [el.split(',') for el in records8]
        # transform to flat list
        flat_list = [item for sublist in rez for item in sublist]
        # remove whitespaces
        flat_list_2 = []
        for el in flat_list:
            flat_list_2.append(" ".join(el.split()))
        flat_list_2 = sorted(flat_list_2)
        # count Countries
        d = {}
        for el in flat_list_2:
            if el not in d:
                d[el] = 1
            else:
                d[el] += 1

        rez = d.items()
        for el in rez:
            book = openpyxl.load_workbook("data/statistics_calculated.xlsx")
            sheet = book.active
            sheet.append(el)
            book.save("data/statistics_calculated.xlsx")

    else:
        #  SELECT column name & subcolumn_name
        start_1 = timeit.timeit()
        get_col_name = """SELECT %s FROM test WHERE id_count=2""" % col_name
        cursor.execute(get_col_name)
        # remove abundant tuple
        col_name_full = [item for t in cursor.fetchall() for item in t]
        end_1 = timeit.timeit()
        print("get_col_name_time", end_1 - start_1)

        get_sub_col_name = """SELECT %s FROM test WHERE id_count=3""" % col_name
        cursor.execute(get_sub_col_name)
        sub_col_name_full = [item for t in cursor.fetchall() for item in t]

        #  Check if column name is not empty.
        if not col_name_full == ['']:
            start_2 = timeit.timeit()
            book = openpyxl.load_workbook("data/statistics_calculated.xlsx")
            sheet = book.active
            sheet.append(col_name_full)
            book.save("data/statistics_calculated.xlsx")
            end_2 = timeit.timeit()
            print("write_col_name_to_file_time", end_2 - start_2)
        if not sub_col_name_full == ['']:
            book = openpyxl.load_workbook("data/statistics_calculated.xlsx")
            sheet = book.active
            sheet.append(sub_col_name_full)
            book.save("data/statistics_calculated.xlsx")

        start_3 = timeit.timeit()
        # SELECT distinct data and GROUP & COUNT values
        query_data = """SELECT DISTINCT(%s) FROM test WHERE %s <> '' AND id_count > 3 """ % (col_name, col_name)
        cursor.execute(query_data)
        global data2
        data2 = cursor.fetchall()
        data2 = [item for t in data2 for item in t]
        data2 = [el.replace("'", "''") for el in data2]
        end_3 = timeit.timeit()
        print("get distinct elements", end_3 - start_3)

        start_7 = timeit.timeit()
        for el in data2:
            start_4 = timeit.timeit()
            cursor.execute(
                """SELECT %s,Count(*) FROM test WHERE %s='%s' GROUP BY %s """ % (col_name, col_name, el, col_name))
            rez = cursor.fetchall()
            end_4 = timeit.timeit()
            print("count values for all elements", end_4 - start_4)

            start_6 = timeit.timeit()
            book = openpyxl.load_workbook("data/statistics_calculated.xlsx")
            sheet = book.active
            sheet.append(rez[0])
            book.save("data/statistics_calculated.xlsx")
            end_6 = timeit.timeit()
            print("write each element to file", end_6 - start_6)
        end_7 = timeit.timeit()
        print("wrote all element", end_7 - start_7)

    # Save to txt

    # with open("data/statistics.txt", "a") as f:
    #     f.write(str(rez))
    #     f.write("\n")
    #     f.close()
