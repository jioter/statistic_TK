import sqlite3

import openpyxl

# import psycopg2
#
# database = psycopg2.connect(database="test", user="user1", password="password", host="localhost", port="5432")
# cursor = database.cursor()
conn = sqlite3.connect('data/statistic_db.db')
cursor = conn.cursor()


def query_filter(col_name, filter_col_1, filter_val_1,
                 filter_col_2, filter_val_2,
                 filter_col_3, filter_val_3,
                 filter_col_4, filter_val_4,
                 filter_col_5, filter_val_5,
                 filter_col_6, filter_val_6):
    """
      Filter data by all combinations in first 6 columns.
      Select poll question and count sum of certain answers to this question.
      :param col_name:
      :return: certain question and sum of answers
      """
    # separate query for column 8
    if col_name == 'col_8':
        get_col_name_8 = """SELECT %s FROM test WHERE id_count=2""" % col_name
        cursor.execute(get_col_name_8)
        col_name_8 = [item for t in cursor.fetchall() for item in t]

        query8 = """SELECT DISTINCT(%s) FROM test 
        WHERE %s <> '' AND id_count > 3 AND %s NOT IN ('-') 
        AND %s=%s AND %s=%s AND %s=%s AND %s=%s AND %s=%s AND %s=%s""" \
                 % (col_name, col_name, col_name,
                    filter_col_1, "'" + filter_val_1 + "'",
                    filter_col_2, "'" + filter_val_2 + "'",
                    filter_col_3, "'" + filter_val_3 + "'",
                    filter_col_4, "'" + filter_val_4 + "'",
                    filter_col_5, "'" + filter_val_5 + "'",
                    filter_col_6, "'" + filter_val_6 + "'")
        cursor.execute(query8)
        records8 = cursor.fetchall()
        records8 = [item for t in records8 for item in t]
        records8 = [el.replace("'", "''") for el in records8]

        # split elements in sublists
        rez = [el.split(',') for el in records8]
        # transform to flat list
        flat_list = [item for sublist in rez for item in sublist]
        # remove whitespaces
        flat_list_2 = []
        for el in flat_list:
            flat_list_2.append(" ".join(el.split()))
        flat_list_2 = sorted(flat_list_2)
        # count Countries
        d = {}
        for el in flat_list_2:
            if el not in d:
                d[el] = 1
            else:
                d[el] += 1

        rez = d.items()
        if not rez:
            pass
        else:
            book = openpyxl.load_workbook("data/statistics_calculated.xlsx")
            sheet = book.active
            sheet.append(col_name_8)
            book.save("data/statistics_calculated.xlsx")

            for el in rez:
                book = openpyxl.load_workbook("data/statistics_calculated.xlsx")
                sheet = book.active
                sheet.append(el)
                book.save("data/statistics_calculated.xlsx")

    else:
        get_col_name = """SELECT %s FROM test WHERE id_count=2""" % col_name
        cursor.execute(get_col_name)
        col_name_full = [item for t in cursor.fetchall() for item in t]

        get_sub_col_name = """SELECT %s FROM test WHERE id_count=3""" % col_name
        cursor.execute(get_sub_col_name)
        sub_col_name_full = [item for t in cursor.fetchall() for item in t]

        query_data = """SELECT DISTINCT(%s) FROM test 
        WHERE %s <> '' AND id_count > 3 
        AND %s=%s AND %s=%s AND %s=%s AND %s=%s AND %s=%s AND %s=%s""" \
                     % (col_name, col_name,
                        filter_col_1, "'" + filter_val_1 + "'",
                        filter_col_2, "'" + filter_val_2 + "'",
                        filter_col_3, "'" + filter_val_3 + "'",
                        filter_col_4, "'" + filter_val_4 + "'",
                        filter_col_5, "'" + filter_val_5 + "'",
                        filter_col_6, "'" + filter_val_6 + "'")
        cursor.execute(query_data)
        global data2
        data2 = cursor.fetchall()
        data2 = [item for t in data2 for item in t]
        data2 = [el.replace("'", "''") for el in data2]

        if not data2:
            pass
        else:
            #  Check if column name is not empty.
            if not col_name_full == ['']:
                book = openpyxl.load_workbook("data/statistics_calculated.xlsx")
                sheet = book.active
                sheet.append(col_name_full)
                book.save("data/statistics_calculated.xlsx")
            if not sub_col_name_full == ['']:
                book = openpyxl.load_workbook("data/statistics_calculated.xlsx")
                sheet = book.active
                sheet.append(sub_col_name_full)
                book.save("data/statistics_calculated.xlsx")

            for el in data2:
                cursor.execute(
                    """SELECT %s,Count(*) FROM test WHERE %s='%s' GROUP BY %s """ % (col_name, col_name, el, col_name))
                rez = cursor.fetchall()
                book = openpyxl.load_workbook("data/statistics_calculated.xlsx")
                sheet = book.active
                sheet.append(rez[0])
                book.save("data/statistics_calculated.xlsx")
